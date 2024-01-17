import openpyxl
import xlsxwriter
 
# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("Bangla-Words.xlsx")
# Define variable to read sheet
dataframe1 = dataframe.active
# Iterate the loop to read the cell values
words = []
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        # print(col[row].value)
        words.append(col[row].value)

# print(dataframe1.max_row)
# print(dataframe1.max_column)

# SORTING
words.sort()

workbook = xlsxwriter.Workbook('Sorted-Bangla-Words.xlsx')
worksheet = workbook.add_worksheet()
row = 0
column = 0
 
# iterating through content list
for item in words :
    # write operation perform
    worksheet.write(row, column, item)
    # incrementing the value of row by one
    # with each iterations.
    row += 1

workbook.close()